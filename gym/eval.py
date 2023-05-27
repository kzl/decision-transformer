import time

import gym
import numpy as np
import torch
import wandb
import json
import openpyxl
import argparse
import pickle
import random
import sys, os, os.path, csv

sys.path.append(os.getcwd())

from decision_transformer.evaluation.evaluate_episodes import evaluate_episode, evaluate_episode_rtg
from decision_transformer.models.decision_transformer import DecisionTransformer
from decision_transformer.models.mlp_bc import MLPBCModel
from decision_transformer.training.act_trainer import ActTrainer
from decision_transformer.training.seq_trainer import SequenceTrainer
from decision_transformer.models.offlinerl.config.algo import cql_config, plas_config, mopo_config, moose_config, bcqd_config, bcq_config, bc_config, crr_config, combo_config, bremen_config, maple_config
from decision_transformer.models.offlinerl.utils.config import parse_config
from decision_transformer.models.offlinerl.algo.modelfree import cql, plas, bcqd, bcq, bc, crr

from decision_transformer.models.offlinerl.algo import algo_select
from decision_transformer.models.offlinerl.data.d4rl import load_d4rl_buffer


def evaluate_episode_cql(
        env,
        state_dim,
        act_dim,
        model,
        max_ep_len=1000,
        device='cuda',
        target_return=None,
        mode='normal',
        state_mean=None,
        state_std=None,
        model_type='cql'
):

    model.eval()
    model.to(device=device)

    state = env.reset()
    # dimOfEnvState = state.shape[0]
    # dimOfInputState = state_mean.shape[0]

    if model_type in ['cqlR', 'cqlR2']:
        rtgMean = state_mean[-1]
        rtgStd = state_std[-1]
        rtgNow = target_return

        state = np.concatenate((state, np.array([(target_return - rtgMean) / (rtgStd + 1e-6)])), )
    # print(state)

    state_mean = torch.from_numpy(state_mean).to(device=device)
    state_std = torch.from_numpy(state_std).to(device=device)

    # we keep all the histories on the device
    # note that the latest action and reward will be "padding"
    states = torch.from_numpy(state).reshape(1, state_dim).to(device=device, dtype=torch.float32)
    actions = torch.zeros((0, act_dim), device=device, dtype=torch.float32)
    rewards = torch.zeros(0, device=device, dtype=torch.float32)
    target_return = torch.tensor(target_return, device=device, dtype=torch.float32)
    sim_states = []

    episode_return, episode_length = 0, 0
    for t in range(max_ep_len):
        # env.render()
        # add padding
        # actions = torch.cat([actions, torch.zeros((1, act_dim), device=device)], dim=0)
        rewards = torch.cat([rewards, torch.zeros(1, device=device)])

        tanh_normal = model((states.to(dtype=torch.float32) - state_mean) / state_std,)
        action = tanh_normal.rsample()

        action = action.detach().cpu().numpy()
        state, reward, done, _ = env.step(action[-1])
        if model_type in ['cqlR', 'cqlR2']:
            if model_type =='cqlR2':
                rtgNow -= reward
            state = np.concatenate((state, np.array([(rtgNow - rtgMean) / (rtgStd + 1e-6)])), )

        cur_state = torch.from_numpy(state).to(device=device, dtype=torch.float32).reshape(1, state_dim)
        states = torch.cat([states, cur_state], dim=0)
        rewards[-1] = reward

        episode_return += reward
        episode_length += 1

        if done:
            break

    return episode_return, episode_length

def discount_cumsum(x, gamma):
    discount_cumsum = np.zeros_like(x)
    discount_cumsum[-1] = x[-1]
    for t in reversed(range(x.shape[0] - 1)):
        discount_cumsum[t] = x[t] + gamma * discount_cumsum[t + 1]
    return discount_cumsum


def experiment(
        exp_prefix,
        variant,
):
    variant["para_path"] = r"saved_para/%s/%s/%s/%s/" % (variant['model_type'], variant["env"], variant["dataset"], variant["mode"])
    config_path = variant["para_path"]+'config.json'

    model_type = variant['model_type']

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

    # 获取目录下所有文件名
    file_names = os.listdir(variant["para_path"])
    # 找出字典序最大的文件名
    # max_file_name = max(file_names)
    # print(file_names)
    using_specified_iter_para = variant['using_specified_iter_para']
    if not using_specified_iter_para:
        max_file_name = [i for i in file_names if "_" in i]
        max_file_name = sorted(max_file_name, key=lambda x: int(x.split('.')[0].split('_')[-1]),reverse=True)
        print(max_file_name)
        para_path = variant["para_path"] + max_file_name[0]
        print("config_path: %s"%(config_path))
        print("para_path: %s" % (para_path))
    else:
        if model_type.startswith("cql"):
            max_file_name = [i for i in file_names if "_" in i]
            max_file_name = sorted(max_file_name, key=lambda x: int(x.split('.')[0].split('_')[1]),reverse=True)
            # print(max_file_name)
            para_path = variant["para_path"] + "iter_%s_actor.pt"%(using_specified_iter_para)
            print("config_path: %s"%(config_path))
            print("para_path: %s" % (para_path))
        else:
            max_file_name = [i for i in file_names if "_" in i]
            max_file_name = sorted(max_file_name, key=lambda x: int(x.split('.')[0].split('_')[-1]),reverse=True)
            # print(max_file_name)
            para_path = variant["para_path"] + "iter_%s.pt"%(using_specified_iter_para)
            print("config_path: %s"%(config_path))
            print("para_path: %s" % (para_path))

    with open(config_path, 'r', encoding='utf-8') as f:
        # 读取json文件内容并转换成字典类型
        data_dict = json.load(f)

    for k,v in data_dict.items():
        if k not in variant.keys():
            variant[k] = v

    # device = variant.get('device', 'cuda')

    log_to_wandb = variant.get('log_to_wandb', False)

    env_name, dataset = variant['env'], variant['dataset']

    group_name = f'{exp_prefix}-{env_name}-{dataset}'
    exp_prefix = f'{group_name}-{random.randint(int(1e5), int(1e6) - 1)}'

    normalized_score_basis = {
        "hopper": [3234.3, -20.272305],
        "halfcheetah": [12135.0, -280.178953],
        "walker2d": [4592.3, 1.629008],
        "ant": [3879.7, -325.6]
    }

    expert_score = normalized_score_basis[env_name][0]
    random_score = normalized_score_basis[env_name][1]

    if env_name == 'hopper':
        env = gym.make('Hopper-v3')
        max_ep_len = 1000
        # env_targets = [3600, 1800]  # evaluation conditioning targets
        env_targets = [3600]  # evaluation conditioning targets
        # env_targets = [i for i in range(2500,3601,100)]  # evaluation conditioning targets
        scale = 1000.  # normalization for rewards/returns
    elif env_name == 'halfcheetah':
        env = gym.make('HalfCheetah-v3')
        max_ep_len = 1000
        # env_targets = [12000, 6000]
        env_targets = [12000]
        scale = 1000.
    elif env_name == 'walker2d':
        env = gym.make('Walker2d-v3')
        max_ep_len = 1000
        # env_targets = [5000, 2500]
        env_targets = [5000]
        scale = 1000.
    elif env_name == 'reacher2d':
        from decision_transformer.envs.reacher_2d import Reacher2dEnv
        env = Reacher2dEnv()
        max_ep_len = 100
        # env_targets = [76, 40]
        env_targets = [76]
        scale = 10.
    else:
        raise NotImplementedError

    if model_type == 'bc':
        env_targets = env_targets[:1]  # since BC ignores target, no need for different evaluations

    state_dim = env.observation_space.shape[0]
    act_dim = env.action_space.shape[0]

    # load dataset
    dataset_path = f'data/{env_name}-{dataset}-v2.pkl'
    with open(dataset_path, 'rb') as f:
        trajectories = pickle.load(f)

    # save all path information into separate lists
    mode = variant.get('mode', 'normal')
    states, traj_lens, returns = [], [], []
    for path in trajectories:
        if mode == 'delayed':  # delayed: all rewards moved to end of trajectory
            path['rewards'][-1] = path['rewards'].sum()
            path['rewards'][:-1] = 0.
        states.append(path['observations'])
        traj_lens.append(len(path['observations']))
        returns.append(path['rewards'].sum())
    traj_lens, returns = np.array(traj_lens), np.array(returns)

    # used for input normalization
    states = np.concatenate(states, axis=0)
    state_mean, state_std = np.mean(states, axis=0), np.std(states, axis=0) + 1e-6

    num_timesteps = sum(traj_lens)

    print('=' * 50)
    print(f'Starting new experiment: {env_name} {dataset}')
    print(f'{len(traj_lens)} trajectories, {num_timesteps} timesteps found')
    print(f'Average return: {np.mean(returns):.2f}, std: {np.std(returns):.2f}')
    print(f'Max return: {np.max(returns):.2f}, min: {np.min(returns):.2f}')
    print('=' * 50)

    K = variant['K']
    batch_size = variant['batch_size']
    num_eval_episodes = variant['num_eval_episodes']
    pct_traj = variant.get('pct_traj', 1.)

    # only train on top pct_traj trajectories (for %BC experiment)
    num_timesteps = max(int(pct_traj * num_timesteps), 1)
    sorted_inds = np.argsort(returns)  # lowest to highest
    num_trajectories = 1
    timesteps = traj_lens[sorted_inds[-1]]
    ind = len(trajectories) - 2
    while ind >= 0 and timesteps + traj_lens[sorted_inds[ind]] <= num_timesteps:
        timesteps += traj_lens[sorted_inds[ind]]
        num_trajectories += 1
        ind -= 1
    sorted_inds = sorted_inds[-num_trajectories:]

    # used to reweight sampling so we sample according to timesteps instead of trajectories
    p_sample = traj_lens[sorted_inds] / sum(traj_lens[sorted_inds])

    def get_batch(batch_size=256, max_len=K):
        '''
        max_len：DT模型输入的一条轨迹的长度
        '''
        batch_inds = np.random.choice(
            np.arange(num_trajectories),
            size=batch_size,
            replace=True,
            p=p_sample,  # reweights so we sample according to timesteps
        )

        s, a, r, d, rtg, timesteps, mask = [], [], [], [], [], [], []
        for i in range(batch_size):
            # 随机取一条轨迹
            traj = trajectories[int(sorted_inds[batch_inds[i]])]
            # 随机取一个轨迹起点
            si = random.randint(0, traj['rewards'].shape[0] - 1)

            # get sequences from dataset
            s.append(traj['observations'][si:si + max_len].reshape(1, -1, state_dim))
            a.append(traj['actions'][si:si + max_len].reshape(1, -1, act_dim))
            r.append(traj['rewards'][si:si + max_len].reshape(1, -1, 1))
            if 'terminals' in traj:
                d.append(traj['terminals'][si:si + max_len].reshape(1, -1))
            else:
                d.append(traj['dones'][si:si + max_len].reshape(1, -1))
            timesteps.append(np.arange(si, si + s[-1].shape[1]).reshape(1, -1))
            # print(timesteps)
            timesteps[-1][timesteps[-1] >= max_ep_len] = max_ep_len-1  # padding cutoff
            rtg.append(discount_cumsum(traj['rewards'][si:], gamma=1.)[:s[-1].shape[1] + 1].reshape(1, -1, 1))

            # import time
            # time.sleep(10)
            if rtg[-1].shape[1] <= s[-1].shape[1]:
                rtg[-1] = np.concatenate([rtg[-1], np.zeros((1, 1, 1))], axis=1)

            # padding and state + reward normalization
            tlen = s[-1].shape[1]
            s[-1] = np.concatenate([np.zeros((1, max_len - tlen, state_dim)), s[-1]], axis=1)
            s[-1] = (s[-1] - state_mean) / state_std
            a[-1] = np.concatenate([np.ones((1, max_len - tlen, act_dim)) * -10., a[-1]], axis=1)
            r[-1] = np.concatenate([np.zeros((1, max_len - tlen, 1)), r[-1]], axis=1)
            d[-1] = np.concatenate([np.ones((1, max_len - tlen)) * 2, d[-1]], axis=1)
            rtg[-1] = np.concatenate([np.zeros((1, max_len - tlen, 1)), rtg[-1]], axis=1) / scale
            timesteps[-1] = np.concatenate([np.zeros((1, max_len - tlen)), timesteps[-1]], axis=1)
            mask.append(np.concatenate([np.zeros((1, max_len - tlen)), np.ones((1, tlen))], axis=1))

        s = torch.from_numpy(np.concatenate(s, axis=0)).to(dtype=torch.float32, device=device)
        a = torch.from_numpy(np.concatenate(a, axis=0)).to(dtype=torch.float32, device=device)
        r = torch.from_numpy(np.concatenate(r, axis=0)).to(dtype=torch.float32, device=device)
        d = torch.from_numpy(np.concatenate(d, axis=0)).to(dtype=torch.long, device=device)
        rtg = torch.from_numpy(np.concatenate(rtg, axis=0)).to(dtype=torch.float32, device=device)
        timesteps = torch.from_numpy(np.concatenate(timesteps, axis=0)).to(dtype=torch.long, device=device)
        mask = torch.from_numpy(np.concatenate(mask, axis=0)).to(device=device)

        # print("dt")

        return s, a, r, d, rtg, timesteps, mask

    def eval_episodes(target_rew, state_mean_, state_std_):
        def fn(model):
            returns, lengths = [], []
            print("eval episode number: %d"%(num_eval_episodes))
            for _ in range(num_eval_episodes):
                with torch.no_grad():
                    if model_type in ('dt', 'qdt'):
                        ret, length = evaluate_episode_rtg(
                            env,
                            state_dim,
                            act_dim,
                            model,
                            max_ep_len=max_ep_len,
                            scale=scale,
                            target_return=target_rew / scale,
                            mode=mode,
                            state_mean=state_mean,
                            state_std=state_std,
                            device=device,
                        )
                    elif model_type=='bc':
                        ret, length = evaluate_episode(
                            env,
                            state_dim,
                            act_dim,
                            model,
                            max_ep_len=max_ep_len,
                            target_return=target_rew / scale,
                            mode=mode,
                            state_mean=state_mean,
                            state_std=state_std,
                            device=device,
                        )
                    elif model_type in ('cql', 'cqlR', 'cqlR2') :
                        ret, length = evaluate_episode_cql(
                            env,
                            state_dim,
                            act_dim,
                            model,
                            max_ep_len=max_ep_len,
                            target_return=target_rew,
                            mode=mode,
                            state_mean=state_mean_,
                            state_std=state_std_,
                            device=device,
                            model_type=model_type
                        )

                returns.append(ret)
                lengths.append(length)
            return {
                f'target_{target_rew}_return_mean': np.mean(returns),
                f'target_{target_rew}_return_std': np.std(returns),
                f'target_{target_rew}_normalized_return_mean': (np.mean(returns)-random_score)/(expert_score-random_score),
                f'target_{target_rew}_normalized_return_std': np.std(returns)/(expert_score-random_score),
                f'target_{target_rew}_length_mean': np.mean(lengths),
                f'target_{target_rew}_length_std': np.std(lengths),
            }

        return fn
    # model_type = 'cql'
    print(model_type)
    if model_type in ('dt', 'qdt'):
        print("why?")
        model = DecisionTransformer(
            state_dim=state_dim,
            act_dim=act_dim,
            max_length=K,
            max_ep_len=max_ep_len,
            hidden_size=variant['embed_dim'],
            n_layer=variant['n_layer'],
            n_head=variant['n_head'],
            n_inner=4 * variant['embed_dim'],
            activation_function=variant['activation_function'],
            n_positions=1024,
            resid_pdrop=variant['dropout'],
            attn_pdrop=variant['dropout'],
        )
        model.load_state_dict(torch.load(para_path, map_location=torch.device(device)))
        model = model.to(device=device)
    elif model_type == 'bc':
        model = MLPBCModel(
            state_dim=state_dim,
            act_dim=act_dim,
            max_length=K,
            hidden_size=variant['embed_dim'],
            n_layer=variant['n_layer'],
        )
        model.load_state_dict(torch.load(para_path, map_location=torch.device(device)))
        model = model.to(device=device)
    elif model_type in ('cql', 'cqlR', 'cqlR2') :

        env_name, dataset = variant['env'], variant['dataset']
        if model_type in ('cqlR', 'cqlR2'):
            state_dim = env.observation_space.shape[0] + 1
        else:
            state_dim = env.observation_space.shape[0]
        act_dim = env.action_space.shape[0]

        variant['state_dim'] = state_dim
        variant['act_dim'] = variant['action_shape'] = act_dim

        algo_init_fn, algo_trainer_obj, algo_config = algo_select(variant, None)
        train_buffer, state_mean, state_std = load_d4rl_buffer(variant)
        algo_init = algo_init_fn
        trainer = algo_trainer_obj(algo_init, algo_config)

        print(env_name, state_dim, act_dim)
        mode = algo_config.get('mode', 'normal')
        device = algo_config["device"]
        device = torch.device(device)

        root_path = os.getcwd()
        trainer.load_pi(root_path + "/saved_para/%s/%s/%s/%s/" % (model_type, env_name, dataset, mode), 300,
                        eval_fn=[eval_episodes(env_targets[0], state_mean, state_std)])
    else:
        raise NotImplementedError

    print("using device: %s" % (device))

    if model_type in ('dt', "qdt"):
        warmup_steps = variant['warmup_steps']
        optimizer = torch.optim.AdamW(
            model.parameters(),
            lr=variant['learning_rate'],
            weight_decay=variant['weight_decay'],
        )
        scheduler = torch.optim.lr_scheduler.LambdaLR(
            optimizer,
            lambda steps: min((steps + 1) / warmup_steps, 1)
        )
        trainer = SequenceTrainer(
            model=model,
            optimizer=optimizer,
            batch_size=batch_size,
            get_batch=get_batch,
            scheduler=scheduler,
            loss_fn=lambda s_hat, a_hat, r_hat, s, a, r: torch.mean((a_hat - a) ** 2),
            eval_fns=[eval_episodes(tar) for tar in env_targets],
        )
    elif model_type == 'bc':
        warmup_steps = variant['warmup_steps']
        optimizer = torch.optim.AdamW(
            model.parameters(),
            lr=variant['learning_rate'],
            weight_decay=variant['weight_decay'],
        )
        scheduler = torch.optim.lr_scheduler.LambdaLR(
            optimizer,
            lambda steps: min((steps + 1) / warmup_steps, 1)
        )
        trainer = ActTrainer(
            model=model,
            optimizer=optimizer,
            batch_size=batch_size,
            get_batch=get_batch,
            scheduler=scheduler,
            loss_fn=lambda s_hat, a_hat, r_hat, s, a, r: torch.mean((a_hat - a) ** 2),
            eval_fns=[eval_episodes(tar) for tar in env_targets],
        )
    # elif model_type in ('cql', 'cqlR', 'cqlR2') :

    #     # get_action = algo_trainer.get_action
    #     # algo_trainer.load_pi(os.getcwd() + "/saved_para/CQL/hopper/expert/normal/", 300, [eval_episodes(0)])
    #     # model = algo_trainer
    #     # trainer = algo_trainer
    #     raise NotImplementedError

    if log_to_wandb:
        wandb.init(
            name=exp_prefix,
            group=group_name,
            project='decision-transformer',
            config=variant
        )
        # wandb.watch(model)  # wandb has some bug
    # for iter in range(variant['max_iters']):
    # if model_type in ('cql', 'cqlR', 'cqlR2') :
    #     # outputs = trainer.eval_para(train_buffer, None, variant=algo_config)
    #     outputs = trainer.eval_para(print_logs=True)
    # else:
    outputs = trainer.eval_para(print_logs=True)

    new_header = "_".join([variant['model_type'] , variant["env"], variant["dataset"], variant["mode"]])
    new_column = []
    for k,v in outputs.items():
        if "length" not in k:
            new_column.append(v)
        else:
            break

    save_path = r"eval_result/"
    if not os.path.exists(save_path):
        os.makedirs(save_path)
    # excel_file = save_path+"%s_%s_%s.xlsx" % (variant["env"], variant["dataset"], variant["mode"])
    excel_file = save_path + "%s.xlsx"%(variant['eval_results_file_name'])
    # excel_file = save_path + "%s_%s_%s_%s.xlsx" % (variant['model_type'], variant["env"], variant["dataset"], variant["mode"])
    if os.path.isfile(excel_file):
        # workbook = pd.read_excel(excel_file)
        workbook = openpyxl.load_workbook(excel_file)
        worksheet = workbook[workbook.sheetnames[0]]
        # 获取现有的表头
        headers = [cell.value for cell in worksheet[1]]
        print("old headers: ", headers)
        worksheet.cell(row=1, column=len(headers) + 1).value = new_header
        print("new headers: ", [cell.value for cell in worksheet[1]])
        for ind, val in enumerate(new_column):
            worksheet.cell(row=ind + 2, column=len(headers) + 1).value = val

        workbook.save(excel_file)
    else:
        # 创建新Excel文件
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # 写入表头
        for ind, col in enumerate(["return_mean", "return_std", "normalized_return_mean", "normalized_return_std"]):
            worksheet.cell(row=ind+2, column=1, value=col)

        # 写入数据
        worksheet.cell(row=1, column=2).value = new_header
        for ind, val in enumerate(new_column):
            worksheet.cell(row=ind+2, column=2, value=val)

        # 保存Excel文件
        workbook.save(excel_file)

    if log_to_wandb:
        wandb.log(outputs)


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--env', type=str, default='halfcheetah')
    parser.add_argument('--dataset', type=str, default='medium')  # medium, medium-replay, medium-expert, expert
    parser.add_argument('--mode', type=str, default='normal')  # normal for standard setting, delayed for sparse
    parser.add_argument('--model_type', type=str, default='cqlR')  # normal for standard setting, delayed for sparse
    parser.add_argument('--num_eval_episodes', type=int, default=3)  # normal for standard setting, delayed for sparse
    parser.add_argument('--K', type=int, default=0)  # normal for standard setting, delayed for sparse
    parser.add_argument('--using_specified_iter_para', type=str, default=300)  # normal for standard setting, delayed for sparse
    parser.add_argument('--eval_results_file_name', type=str, default='eval_results_%s'%(str(time.time())))  # normal for standard setting, delayed for sparse
    parser.add_argument('--device', type=str, default='cpu')

    args = parser.parse_args()

    experiment('gym-experiment', variant=vars(args))

