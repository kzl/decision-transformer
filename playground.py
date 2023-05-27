import openpyxl
#
# # 定义字典数据
# data_dict = {
#     'row1': {'column1': 1, 'column2': 2, 'column3': 3},
#     'row2': {'column1': 4, 'column2': 5, 'column3': 6},
#     'row3': {'column1': 7, 'column2': 8, 'column3': 9}
# }
#
# # 创建新Excel文件
# workbook = openpyxl.Workbook()
# worksheet = workbook.active
#
# # 写入表头
# columns = list(data_dict['row1'].keys())
# for i, col in enumerate(columns):
#     worksheet.cell(row=1, column=i+1, value=col)
#
# # 写入数据
# for i, (row, data) in enumerate(data_dict.items()):
#     worksheet.cell(row=i+2, column=2, value=row)
#     for j, (col, val) in enumerate(data.items()):
#         worksheet.cell(row=i+2, column=j+2, value=val)
#
# # 保存Excel文件
# workbook.save('example.xlsx')



excel_file = 'example.xlsx'
workbook = openpyxl.load_workbook(excel_file)
# worksheets = workbook.sheetnames[0]
# print(worksheets)
worksheet = workbook[workbook.sheetnames[0]]
# 获取现有的表头
headers = [cell.value for cell in worksheet[1]]

print(headers)